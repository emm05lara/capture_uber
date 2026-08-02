"""
Generación de reportes .xlsx para la interfaz operativa (Fase 8).

El módulo se organiza en tres bloques, en este orden:
  1. Helpers y estilos comunes (encabezados, anchos de columna, formatos,
     sanitización de nombres de archivo, respuesta HTTP).
  2. Obtención de datos: pequeñas funciones que reciben querysets/objetos ya
     resueltos por las vistas (o los selectores de `dashboard_queries`) y
     los "aplanan" a filas simples, haciendo como máximo un puñado de
     consultas adicionales por lote (nunca una por fila).
  3. Construcción de libros: arman el `Workbook` de openpyxl a partir de
     esas filas.

Nada de esto toca disco: todo se serializa a `BytesIO` y se entrega como
`HttpResponse` de descarga.
"""
import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from actores.models import Conductor
from dispositivos.models import AsignacionTag, InstalacionGps
from operacion.models import AsignacionVehiculo
from .models import Vehiculo

# ---------------------------------------------------------------------------
# 1. Helpers y estilos comunes
# ---------------------------------------------------------------------------

CONTENT_TYPE_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
FORMATO_FECHA = "dd/mm/yyyy"
FORMATO_MONEDA = "$#,##0.00"
_RELLENO_ENCABEZADO = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_FUENTE_ENCABEZADO = Font(color="FFFFFF", bold=True)
_ESTATUS_VEHICULO_DISPLAY = dict(Vehiculo.EstatusUnidad.choices)
_ESTATUS_CONDUCTOR_DISPLAY = dict(Conductor.Estatus.choices)
_TIPO_LICENCIA_DISPLAY = dict(Conductor.TipoLicencia.choices)


def sanitizar_nombre_archivo(nombre):
    """Convierte un texto libre (número interno, placas, etc.) en un nombre
    de archivo seguro para Content-Disposition: sin espacios, acentos ni
    separadores de ruta."""
    nombre = (nombre or "archivo").strip()
    nombre = re.sub(r"[^A-Za-z0-9_-]+", "_", nombre)
    nombre = nombre.strip("_")
    return nombre or "archivo"


def _escribir_encabezados(ws, encabezados):
    ws.append(encabezados)
    for col_idx in range(1, len(encabezados) + 1):
        celda = ws.cell(row=1, column=col_idx)
        celda.font = _FUENTE_ENCABEZADO
        celda.fill = _RELLENO_ENCABEZADO
        celda.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(encabezados))}1"


def _autoajustar_columnas(ws, num_columnas):
    for idx in range(1, num_columnas + 1):
        letra = get_column_letter(idx)
        max_len = 8
        for celda in ws[letra]:
            if celda.value is not None:
                max_len = max(max_len, len(str(celda.value)))
        ws.column_dimensions[letra].width = min(max_len + 2, 45)


def _valor_para_celda(valor):
    """Excel/openpyxl no admite `datetime` con zona horaria. Los campos de
    fecha del proyecto son `DateField` (naive) casi siempre; el único caso
    con hora (p. ej. `Observacion.fecha_registro`) es un `DateTimeField`
    con `USE_TZ=True`, así que se convierte a hora local sin tzinfo aquí,
    en un solo lugar, en vez de en cada llamador."""
    if isinstance(valor, datetime) and valor.tzinfo is not None:
        return timezone.localtime(valor).replace(tzinfo=None)
    return valor


def agregar_hoja(wb, titulo, encabezados, filas, primera_hoja=False):
    """Crea una hoja, escribe encabezados con estilo + autofiltro + primera
    fila congelada, y cada fila de datos aplicando formato de fecha/moneda
    según el tipo real de cada valor. Si `filas` está vacío, agrega una
    única fila "Sin registros" para que el archivo nunca quede sin
    contenido bajo los encabezados."""
    ws = wb.active if primera_hoja else wb.create_sheet()
    ws.title = titulo[:31]  # límite de Excel para nombres de hoja

    _escribir_encabezados(ws, encabezados)

    if not filas:
        ws.append(["Sin registros"] + [""] * (len(encabezados) - 1))
    else:
        for fila in filas:
            fila = [_valor_para_celda(v) for v in fila]
            ws.append(fila)
            fila_num = ws.max_row
            for col_idx, valor in enumerate(fila, start=1):
                if isinstance(valor, (date, datetime)):
                    ws.cell(row=fila_num, column=col_idx).number_format = FORMATO_FECHA
                elif isinstance(valor, Decimal):
                    ws.cell(row=fila_num, column=col_idx).number_format = FORMATO_MONEDA

    _autoajustar_columnas(ws, len(encabezados))
    return ws


def respuesta_excel(wb, nombre_archivo):
    """Serializa el libro en memoria (sin escribir en disco) y arma la
    respuesta HTTP de descarga con los encabezados correctos."""
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)

    base, _, _ext = nombre_archivo.rpartition(".")
    nombre_final = sanitizar_nombre_archivo(base or nombre_archivo) + ".xlsx"

    response = HttpResponse(buffer.getvalue(), content_type=CONTENT_TYPE_XLSX)
    response["Content-Disposition"] = f'attachment; filename="{nombre_final}"'
    return response


# ---------------------------------------------------------------------------
# 2. Obtención de datos (aplana relaciones ya resueltas, sin N+1)
# ---------------------------------------------------------------------------

def _filas_vehiculos(fichas):
    """A partir de un queryset de `VwFichaVehiculo` ya filtrado (mismos
    filtros que el listado), agrega en 3 consultas adicionales como máximo
    —independientemente del número de vehículos— el teléfono del conductor
    actual y el GPS/TAG actualmente instalados/asignados, que la vista de
    BD no expone."""
    fichas = list(fichas)
    ids = [f.pk for f in fichas]

    conductor_por_vehiculo = {
        a.vehiculo_id: a
        for a in AsignacionVehiculo.objects.filter(
            vehiculo_id__in=ids, fecha_fin__isnull=True
        ).select_related("conductor")
    }
    gps_por_vehiculo = {
        i.vehiculo_id: i
        for i in InstalacionGps.objects.filter(
            vehiculo_id__in=ids, fecha_retiro__isnull=True
        ).select_related("gps")
    }
    tag_por_vehiculo = {
        a.vehiculo_id: a
        for a in AsignacionTag.objects.filter(
            vehiculo_id__in=ids, fecha_fin__isnull=True
        ).select_related("tag")
    }

    filas = []
    for f in fichas:
        asignacion = conductor_por_vehiculo.get(f.pk)
        instalacion_gps = gps_por_vehiculo.get(f.pk)
        asignacion_tag = tag_por_vehiculo.get(f.pk)
        filas.append((
            f.numero_interno or "",
            f.numero_serie,
            f.nombre_marca,
            f.nombre_modelo_comercial,
            f.anio_modelo,
            f.nombre_color,
            f.placas_actuales or "",
            f.entidad_emplacamiento_actual or "",
            _ESTATUS_VEHICULO_DISPLAY.get(f.estatus_unidad, f.estatus_unidad),
            f.conductor or "",
            asignacion.conductor.telefono if asignacion and asignacion.conductor.telefono else "",
            f.numero_poliza or "",
            f.vigencia_poliza,
            f.semestre_verificacion or "",
            f.fecha_limite_verificacion,
            f.vigencia_tarjeta_circulacion,
            instalacion_gps.gps.imei if instalacion_gps else "",
            asignacion_tag.tag.codigo_tag if asignacion_tag else "",
            f.monto_adeudos_pendientes or Decimal("0"),
            f.semaforo_documental or "",
        ))
    return filas


def _placas_actuales_prefetchadas(vehiculo):
    """Igual que `Vehiculo.placas_actuales`, pero leyendo el emplacamiento
    vigente desde un `to_attr` ya prefetcheado (`emplacamiento_actual_list`)
    en vez de disparar una consulta nueva por vehículo."""
    lista = getattr(vehiculo, "emplacamiento_actual_list", None)
    if lista is None:
        return vehiculo.placas_actuales or ""
    return lista[0].placas if lista else ""


def _filas_conductores(conductores):
    """`conductores` debe venir con el prefetch de asignación activa
    (`to_attr="asignacion_activa"`, incluyendo el emplacamiento vigente del
    vehículo como `emplacamiento_actual_list`) y la anotación
    `num_referencias`, tal como los arma
    `actores.views._conductores_filtrados`."""
    filas = []
    for c in conductores:
        asignacion = c.asignacion_activa[0] if getattr(c, "asignacion_activa", None) else None
        filas.append((
            c.nombre_completo,
            c.telefono or "",
            c.curp or "",
            c.numero_licencia or "",
            _TIPO_LICENCIA_DISPLAY.get(c.tipo_licencia, c.tipo_licencia or ""),
            c.fecha_vencimiento_licencia,
            _ESTATUS_CONDUCTOR_DISPLAY.get(c.estatus_conductor, c.estatus_conductor),
            (asignacion.vehiculo.numero_interno or asignacion.vehiculo.numero_serie) if asignacion else "",
            _placas_actuales_prefetchadas(asignacion.vehiculo) if asignacion else "",
            asignacion.fecha_inicio if asignacion else None,
            c.num_referencias if getattr(c, "num_referencias", None) is not None else c.referencias.count(),
        ))
    return filas


# ---------------------------------------------------------------------------
# 3. Construcción de libros
# ---------------------------------------------------------------------------

def construir_libro_vehiculos(fichas):
    wb = Workbook()
    encabezados = [
        "Número interno", "VIN / Número de serie", "Marca", "Modelo", "Año",
        "Color", "Placas actuales", "Entidad federativa", "Estatus",
        "Conductor actual", "Teléfono conductor", "Póliza vigente",
        "Vencimiento póliza", "Verificación actual", "Fecha límite verificación",
        "Vigencia tarjeta circulación", "GPS actual", "TAG actual",
        "Adeudos pendientes", "Semáforo documental",
    ]
    agregar_hoja(wb, "Vehículos", encabezados, _filas_vehiculos(fichas), primera_hoja=True)
    return wb


def construir_libro_conductores(conductores):
    wb = Workbook()
    encabezados = [
        "Nombre", "Teléfono", "CURP", "Número de licencia", "Tipo de licencia",
        "Vigencia de licencia", "Estatus", "Vehículo actual", "Placas actuales",
        "Inicio de asignación", "Referencias personales",
    ]
    agregar_hoja(wb, "Conductores", encabezados, _filas_conductores(conductores), primera_hoja=True)
    return wb


def construir_libro_ficha_vehiculo(vehiculo, datos):
    """`datos` es el dict de contexto que también usa `detalle_vehiculo`
    (ver `vehiculos.views._datos_detalle_vehiculo`), para no duplicar las
    consultas que arma esa vista."""
    wb = Workbook()

    ficha = datos.get("ficha")
    resumen = [(
        vehiculo.numero_interno or "",
        vehiculo.numero_serie,
        f"{vehiculo.modelo_vehiculo.marca} {vehiculo.modelo_vehiculo.nombre_modelo_comercial}",
        vehiculo.anio_modelo,
        vehiculo.color.nombre_color,
        _ESTATUS_VEHICULO_DISPLAY.get(vehiculo.estatus_unidad, vehiculo.estatus_unidad),
        (ficha.placas_actuales if ficha else None) or "",
        datos["asignacion_actual"].conductor.nombre_completo if datos["asignacion_actual"] else "",
        (ficha.semaforo_documental if ficha else None) or "",
        datos["gps_actual"].gps.imei if datos["gps_actual"] else "",
        datos["tag_actual"].tag.codigo_tag if datos["tag_actual"] else "",
    )]
    agregar_hoja(wb, "Resumen", [
        "Número interno", "VIN", "Marca y modelo", "Año", "Color", "Estatus",
        "Placas actuales", "Conductor actual", "Semáforo documental",
        "GPS actual", "TAG actual",
    ], resumen, primera_hoja=True)

    agregar_hoja(wb, "Emplacamientos", ["Placas", "Entidad", "Inicio", "Fin", "Estado"], [
        (e.placas, str(e.entidad_federativa), e.fecha_inicio, e.fecha_fin, "Actual" if e.es_actual else "Finalizado")
        for e in datos["emplacamientos"]
    ])

    agregar_hoja(wb, "Asignaciones", [
        "Conductor", "Inicio", "Fin", "Plataforma", "Socio", "Estatus",
    ], [
        (
            a.conductor.nombre_completo, a.fecha_inicio, a.fecha_fin,
            str(a.plataforma) if a.plataforma else "", str(a.socio) if a.socio else "",
            a.estatus_asignacion,
        )
        for a in datos["asignaciones"]
    ])

    agregar_hoja(wb, "Pólizas", [
        "Aseguradora", "Titular", "Número", "Inicio vigencia", "Fin vigencia", "Importe", "Estatus",
    ], [
        (
            str(p.aseguradora), str(p.titular_poliza) if p.titular_poliza else "", p.numero_poliza,
            p.fecha_vigencia_inicio, p.fecha_vigencia_fin, p.importe_prima or Decimal("0"),
            "Vigente" if p.vigente else "Vencida",
        )
        for p in datos["polizas"]
    ])

    agregar_hoja(wb, "Verificaciones", ["Semestre", "Última verificación", "Fecha límite", "Estatus"], [
        (v.semestre, v.fecha_ultima_verificacion, v.fecha_limite_verificacion, "Vencida" if v.vencida else "Vigente")
        for v in datos["verificaciones"]
    ])

    agregar_hoja(wb, "Tarjetas", ["Emisión", "Vigencia hasta", "Estatus"], [
        (t.fecha_emision, t.fecha_vigencia_fin, "Vigente" if t.vigente else "Vencida")
        for t in datos["tarjetas"]
    ])

    agregar_hoja(wb, "Tenencias", ["Año fiscal", "Estatus", "Monto", "Fecha de pago"], [
        (t.anio_fiscal, t.get_estatus_tenencia_display(), t.monto_tenencia or Decimal("0"), t.fecha_pago)
        for t in datos["tenencias"]
    ])

    agregar_hoja(wb, "Adeudos", ["Tipo", "Monto", "Estatus", "Fecha de consulta", "Observación"], [
        (a.tipo_adeudo, a.monto_adeudo or Decimal("0"), a.get_estatus_adeudo_display(), a.fecha_consulta, a.observacion_adeudo or "")
        for a in datos["adeudos"]
    ])

    agregar_hoja(wb, "Observaciones", ["Tipo", "Observación", "Fecha", "Autor"], [
        (
            o.get_tipo_observacion_display(), o.texto_observacion, o.fecha_registro,
            str(o.autor_registro) if o.autor_registro else "",
        )
        for o in datos["observaciones"]
    ])

    agregar_hoja(wb, "GPS", ["Dispositivo", "IMEI", "Instalación", "Retiro", "Estatus"], [
        (
            i.gps.numero_gps or "", i.gps.imei, i.fecha_instalacion, i.fecha_retiro,
            "Instalado" if i.es_actual else "Retirado",
        )
        for i in datos["instalaciones_gps"]
    ])

    agregar_hoja(wb, "TAG", ["Código", "Inicio", "Fin", "Estatus"], [
        (a.tag.codigo_tag, a.fecha_inicio, a.fecha_fin, "Activo" if a.es_actual else "Inactivo")
        for a in datos["asignaciones_tag"]
    ])

    return wb


def construir_libro_dashboard(vencimientos, faltantes, adeudos):
    wb = Workbook()

    agregar_hoja(wb, "Vencimientos", [
        "Severidad", "Tipo", "Vehículo o conductor", "Placas", "Fecha", "Días", "Referencia",
    ], [
        (v["severidad_label"], v["tipo_label"], v["entidad"], v["placas"] or "", v["fecha"], v["dias"], v["url"])
        for v in vencimientos
    ], primera_hoja=True)

    agregar_hoja(wb, "Información faltante", ["Vehículo", "Placas", "Faltantes detectados", "Estatus"], faltantes)

    agregar_hoja(wb, "Adeudos pendientes", ["Vehículo", "Placas", "Tipo", "Monto", "Fecha de consulta"], [
        (
            a.vehiculo.numero_interno or a.vehiculo.numero_serie,
            a.vehiculo.placas_actuales or "",
            a.tipo_adeudo,
            a.monto_adeudo or Decimal("0"),
            a.fecha_consulta,
        )
        for a in adeudos
    ])

    return wb
