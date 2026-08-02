"""Pruebas de la Fase 8: exportación a Excel."""
from datetime import date, timedelta
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.test.utils import CaptureQueriesContext
from django.db import connection
from django.urls import reverse
from openpyxl import load_workbook

from actores.models import Aseguradora, Conductor
from catalogos.models import Color, Marca, ModeloVehiculo
from dispositivos.models import AsignacionTag, DispositivoGps, InstalacionGps, TagTelepeaje
from operacion.models import AsignacionVehiculo
from .exporters import sanitizar_nombre_archivo
from .models import AdeudoVehicular, Observacion, PolizaSeguro, TarjetaCirculacion, Tenencia, Vehiculo, VerificacionVehicular

User = get_user_model()

CONTENT_TYPE_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def crear_vehiculo(numero_serie, numero_interno=None, estatus=Vehiculo.EstatusUnidad.ACTIVA):
    marca = Marca.objects.create(nombre_marca=f"Marca {numero_serie}")
    modelo = ModeloVehiculo.objects.create(marca=marca, nombre_modelo_comercial="Modelo X")
    color = Color.objects.create(nombre_color=f"Color {numero_serie}")
    return Vehiculo.objects.create(
        numero_serie=numero_serie,
        numero_interno=numero_interno,
        modelo_vehiculo=modelo,
        color=color,
        anio_modelo=2023,
        estatus_unidad=estatus,
    )


def cargar_libro(response):
    return load_workbook(BytesIO(response.content))


class SanitizarNombreArchivoTests(TestCase):
    def test_quita_caracteres_invalidos(self):
        self.assertEqual(sanitizar_nombre_archivo("V-001"), "V-001")
        self.assertEqual(sanitizar_nombre_archivo("ABC 123 / raro?.xlsx"), "ABC_123_raro_xlsx")

    def test_nombre_vacio_regresa_archivo_por_defecto(self):
        self.assertEqual(sanitizar_nombre_archivo(""), "archivo")
        self.assertEqual(sanitizar_nombre_archivo(None), "archivo")

    def test_no_deja_barras_de_ruta(self):
        resultado = sanitizar_nombre_archivo("../../etc/passwd")
        self.assertNotIn("/", resultado)
        self.assertNotIn("..", resultado)


class ExportarListaVehiculosTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="expveh1", password="clave-segura-123")
        self.client.login(username="expveh1", password="clave-segura-123")

    def test_requiere_login(self):
        self.client.logout()
        resp = self.client.get(reverse("vehiculos:exportar_lista"))
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/login", resp.url)

    def test_status_y_content_type(self):
        crear_vehiculo("VEXP0000000001")
        resp = self.client.get(reverse("vehiculos:exportar_lista"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], CONTENT_TYPE_XLSX)
        self.assertIn("attachment", resp["Content-Disposition"])
        self.assertIn(".xlsx", resp["Content-Disposition"])

    def test_archivo_valido_y_encabezados_esperados(self):
        crear_vehiculo("VEXP0000000002", numero_interno="V-EXP-02")
        resp = self.client.get(reverse("vehiculos:exportar_lista"))
        wb = cargar_libro(resp)
        self.assertIn("Vehículos", wb.sheetnames)
        ws = wb["Vehículos"]
        encabezados = [c.value for c in ws[1]]
        self.assertEqual(encabezados[0], "Número interno")
        self.assertIn("Semáforo documental", encabezados)
        self.assertIn("Adeudos pendientes", encabezados)
        # primera fila congelada + autofiltro habilitado
        self.assertEqual(ws.freeze_panes, "A2")
        self.assertIsNotNone(ws.auto_filter.ref)

    def test_respeta_filtro_de_alerta(self):
        con_conductor = crear_vehiculo("VEXP0000000010", numero_interno="V-CON")
        sin_conductor = crear_vehiculo("VEXP0000000011", numero_interno="V-SIN")
        conductor = Conductor.objects.create(nombre_completo="Conductor Export")
        AsignacionVehiculo.objects.create(vehiculo=con_conductor, conductor=conductor, fecha_fin=None)

        resp = self.client.get(reverse("vehiculos:exportar_lista"), {"alerta": "sin_conductor"})
        wb = cargar_libro(resp)
        ws = wb["Vehículos"]
        vins = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
        self.assertIn("VEXP0000000011", vins)
        self.assertNotIn("VEXP0000000010", vins)

    def test_exporta_todos_los_resultados_no_solo_una_pagina(self):
        # El listado pagina de 25 en 25; con 30 vehículos que cumplen el
        # filtro, el Excel debe traer los 30, no solo los primeros 25.
        for i in range(30):
            crear_vehiculo(f"VEXPPAG{i:09d}")
        resp = self.client.get(reverse("vehiculos:exportar_lista"), {"q": "VEXPPAG"})
        wb = cargar_libro(resp)
        ws = wb["Vehículos"]
        filas = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(filas), 30)

    def test_numero_de_consultas_estable(self):
        for i in range(3):
            crear_vehiculo(f"VEXPQ00000000{i}")
        with CaptureQueriesContext(connection) as ctx:
            resp = self.client.get(reverse("vehiculos:exportar_lista"))
        self.assertEqual(resp.status_code, 200)
        pocas = len(ctx.captured_queries)

        for i in range(3, 15):
            crear_vehiculo(f"VEXPQ00000000{i}")
        with CaptureQueriesContext(connection) as ctx2:
            resp2 = self.client.get(reverse("vehiculos:exportar_lista"))
        self.assertEqual(resp2.status_code, 200)
        muchas = len(ctx2.captured_queries)

        self.assertEqual(pocas, muchas)


class ExportarDetalleVehiculoTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="expveh2", password="clave-segura-123")
        self.client.login(username="expveh2", password="clave-segura-123")
        self.aseguradora = Aseguradora.objects.create(nombre_organizacion="Aseguradora Export")

    def test_requiere_login(self):
        v = crear_vehiculo("VEXPDET000001")
        self.client.logout()
        resp = self.client.get(reverse("vehiculos:exportar_detalle", args=[v.pk]))
        self.assertEqual(resp.status_code, 302)

    def test_vehiculo_inexistente_da_404(self):
        resp = self.client.get(reverse("vehiculos:exportar_detalle", args=[999999]))
        self.assertEqual(resp.status_code, 404)

    def test_hojas_esperadas(self):
        v = crear_vehiculo("VEXPDET000002", numero_interno="V-DET-02")
        resp = self.client.get(reverse("vehiculos:exportar_detalle", args=[v.pk]))
        self.assertEqual(resp.status_code, 200)
        wb = cargar_libro(resp)
        esperadas = [
            "Resumen", "Emplacamientos", "Asignaciones", "Pólizas", "Verificaciones",
            "Tarjetas", "Tenencias", "Adeudos", "Observaciones", "GPS", "TAG",
        ]
        self.assertEqual(wb.sheetnames, esperadas)

    def test_estado_vacio_no_rompe_archivo(self):
        v = crear_vehiculo("VEXPDET000003")
        resp = self.client.get(reverse("vehiculos:exportar_detalle", args=[v.pk]))
        self.assertEqual(resp.status_code, 200)
        wb = cargar_libro(resp)
        ws = wb["Pólizas"]
        fila = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(fila[0], "Sin registros")

    def test_incluye_datos_actuales_e_historicos(self):
        v = crear_vehiculo("VEXPDET000004")
        PolizaSeguro.objects.create(
            vehiculo=v, aseguradora=self.aseguradora, numero_poliza="POL-HIST-1",
            fecha_vigencia_fin=date(2020, 1, 1),
        )
        PolizaSeguro.objects.create(
            vehiculo=v, aseguradora=self.aseguradora, numero_poliza="POL-ACTUAL",
            fecha_vigencia_fin=date(2030, 1, 1),
        )
        resp = self.client.get(reverse("vehiculos:exportar_detalle", args=[v.pk]))
        wb = cargar_libro(resp)
        ws = wb["Pólizas"]
        numeros = [row[2] for row in ws.iter_rows(min_row=2, values_only=True)]
        self.assertIn("POL-HIST-1", numeros)
        self.assertIn("POL-ACTUAL", numeros)

    def test_nombre_de_archivo_usa_numero_interno(self):
        v = crear_vehiculo("VEXPDET000005", numero_interno="V-007")
        resp = self.client.get(reverse("vehiculos:exportar_detalle", args=[v.pk]))
        self.assertIn("ficha_V-007", resp["Content-Disposition"])


class ExportarConductoresTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="expcond1", password="clave-segura-123")
        self.client.login(username="expcond1", password="clave-segura-123")

    def test_requiere_login(self):
        self.client.logout()
        resp = self.client.get(reverse("actores:conductores_exportar"))
        self.assertEqual(resp.status_code, 302)

    def test_respeta_filtros(self):
        Conductor.objects.create(
            nombre_completo="Con Licencia Vencida",
            fecha_vencimiento_licencia=date(2020, 1, 1),
        )
        Conductor.objects.create(nombre_completo="Sin Problemas")
        resp = self.client.get(reverse("actores:conductores_exportar"), {"alerta": "licencia_vencida"})
        wb = cargar_libro(resp)
        ws = wb["Conductores"]
        nombres = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
        self.assertIn("Con Licencia Vencida", nombres)
        self.assertNotIn("Sin Problemas", nombres)

    def test_no_expone_referencias_personales(self):
        from actores.models import ReferenciaConductor

        conductor = Conductor.objects.create(nombre_completo="Con Referencia")
        ReferenciaConductor.objects.create(
            conductor=conductor, nombre="Contacto Secreto", domicilio="Calle Privada 123",
            telefono_contacto="555-000-0000", parentesco="MADRE",
        )
        resp = self.client.get(reverse("actores:conductores_exportar"))
        wb = cargar_libro(resp)
        ws = wb["Conductores"]
        contenido = "\n".join(
            str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None
        )
        self.assertNotIn("Contacto Secreto", contenido)
        self.assertNotIn("Calle Privada 123", contenido)
        self.assertNotIn("555-000-0000", contenido)

    def test_columnas_principales_presentes(self):
        resp = self.client.get(reverse("actores:conductores_exportar"))
        wb = cargar_libro(resp)
        ws = wb["Conductores"]
        encabezados = [c.value for c in ws[1]]
        self.assertEqual(encabezados, [
            "Nombre", "Teléfono", "CURP", "Número de licencia", "Tipo de licencia",
            "Vigencia de licencia", "Estatus", "Vehículo actual", "Placas actuales",
            "Inicio de asignación", "Referencias personales",
        ])


class ExportarDashboardTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="expdash1", password="clave-segura-123")
        self.client.login(username="expdash1", password="clave-segura-123")
        self.aseguradora = Aseguradora.objects.create(nombre_organizacion="Aseguradora Dash Export")

    def test_requiere_login(self):
        self.client.logout()
        resp = self.client.get(reverse("vehiculos:dashboard_exportar"))
        self.assertEqual(resp.status_code, 302)

    def test_hojas_esperadas(self):
        resp = self.client.get(reverse("vehiculos:dashboard_exportar"))
        self.assertEqual(resp.status_code, 200)
        wb = cargar_libro(resp)
        self.assertEqual(wb.sheetnames, ["Vencimientos", "Información faltante", "Adeudos pendientes"])

    def test_respeta_severidad_y_horizonte(self):
        from django.utils.timezone import localdate

        hoy = localdate()
        vencida = crear_vehiculo("VEXPDASH00001")
        proxima = crear_vehiculo("VEXPDASH00002")
        PolizaSeguro.objects.create(
            vehiculo=vencida, aseguradora=self.aseguradora, numero_poliza="POL-DASH-V",
            fecha_vigencia_fin=hoy - timedelta(days=5),
        )
        PolizaSeguro.objects.create(
            vehiculo=proxima, aseguradora=self.aseguradora, numero_poliza="POL-DASH-P",
            fecha_vigencia_fin=hoy + timedelta(days=10),
        )
        resp = self.client.get(reverse("vehiculos:dashboard_exportar"), {"severidad": "VENCIDO"})
        wb = cargar_libro(resp)
        ws = wb["Vencimientos"]
        vehiculos_listados = [row[2] for row in ws.iter_rows(min_row=2, values_only=True)]
        self.assertIn("VEXPDASH00001", vehiculos_listados)
        self.assertNotIn("VEXPDASH00002", vehiculos_listados)

    def test_monto_adeudos_correcto(self):
        v = crear_vehiculo("VEXPDASH00003")
        AdeudoVehicular.objects.create(vehiculo=v, tipo_adeudo="Multa", estatus_adeudo="PENDIENTE", monto_adeudo=1500)
        resp = self.client.get(reverse("vehiculos:dashboard_exportar"))
        wb = cargar_libro(resp)
        ws = wb["Adeudos pendientes"]
        filas = list(ws.iter_rows(min_row=2, values_only=True))
        montos = [f[3] for f in filas if f[0] == "VEXPDASH00003"]
        self.assertEqual(montos, [1500])

    def test_informacion_faltante_incluye_estatus(self):
        crear_vehiculo("VEXPDASH00004")
        resp = self.client.get(reverse("vehiculos:dashboard_exportar"))
        wb = cargar_libro(resp)
        ws = wb["Información faltante"]
        encabezados = [c.value for c in ws[1]]
        self.assertEqual(encabezados, ["Vehículo", "Placas", "Faltantes detectados", "Estatus"])
